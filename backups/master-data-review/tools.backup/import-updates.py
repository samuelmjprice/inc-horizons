#!/usr/bin/env python3
"""Draft approved tracker rows into a non-destructive JSON review file.

This helper reads HORIZONS_Website_Update_Tracker.xlsx and writes approved rows
to data/output/approved-updates.json. It does not overwrite content.json.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "data" / "input" / "HORIZONS_Website_Update_Tracker.xlsx"
FALLBACK_INPUT = ROOT / "HORIZONS_Website_Update_Tracker.xlsx"
OUTPUT = ROOT / "data" / "output" / "approved-updates.json"
READY_STATUSES = {"approved", "ready for codex"}


def row_status(row: dict[str, object]) -> str:
    for key in ("Status", "Website Update Status", "New Status", "Current Status"):
        value = row.get(key)
        if value:
            return str(value).strip()
    return ""


def read_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True)
    approved: list[dict[str, object]] = []

    for sheet in workbook.worksheets:
        if sheet.title == "START HERE":
            continue
        headers = [cell.value for cell in sheet[1]]
        if not any(headers):
            continue
        for cells in sheet.iter_rows(min_row=2, values_only=True):
            row = {
                str(header): value
                for header, value in zip(headers, cells)
                if header and value not in (None, "")
            }
            if not row:
                continue
            status = row_status(row).lower()
            if status in READY_STATUSES:
                approved.append({"sourceTab": sheet.title, "row": row})

    return approved


def main() -> None:
    input_path = DEFAULT_INPUT if DEFAULT_INPUT.exists() else FALLBACK_INPUT
    if not input_path.exists():
        raise SystemExit(
            "No tracker found. Place HORIZONS_Website_Update_Tracker.xlsx in "
            "data/input/ or the project root."
        )

    approved = read_rows(input_path)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps({
        "source": str(input_path.relative_to(ROOT)),
        "note": "Review these approved rows manually before updating content.json.",
        "approvedRows": approved,
    }, indent=2), encoding="utf-8")
    print(f"Wrote {len(approved)} approved rows to {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
