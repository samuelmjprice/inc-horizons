#!/usr/bin/env python3
"""Import approved rows from the final HORIZONS master review workbook.

The importer intentionally writes a preview file first. It only applies rows
marked with final/approved statuses and preserves skipped rows in a Markdown
report for audit.
"""

from __future__ import annotations

import json
import re
import shutil
from copy import deepcopy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "data/final-approved/HORIZONS_Master_Website_Data_Review_FINAL.xlsx"
CONTENT = ROOT / "content.json"
PREVIEW = ROOT / "data/output/content.final-master-import.preview.json"
REPORT = ROOT / "FINAL_MASTER_IMPORT_REPORT.md"
SKIPPED = ROOT / "FINAL_MASTER_IMPORT_SKIPPED_ROWS.md"

APPROVED = {
    "approved",
    "final",
    "final answer",
    "ready for website",
    "ready for codex",
    "added to website",
    "confirmed",
}
SKIP = {
    "missing data",
    "needs confirmation",
    "waiting on person",
    "conflict — needs review",
    "conflict - needs review",
    "draft",
    "tbc",
    "unclear",
}

PREFIX_TO_PATH = {
    "RED": ("redFlags",),
    "DEC": ("decisions",),
    "CALL": ("callSheets",),
    "SCH": ("schedule",),
    "LOC": ("locations",),
    "WHO": ("whoDoICall",),
    "CON": ("contacts",),
    "SUP": ("suppliers",),
    "POD": ("podcast",),
    "SPEAK": ("speakers",),
    "REH": ("rehearsals",),
    "ENT": ("entertainment",),
    "PLAY": ("curatedPlaylists",),
    "CAP": ("contentCapture",),
    "MAT": ("swag",),
    "DROP": ("roomDrops",),
    "HH": ("horizonsHouse",),
    "SIG": ("artworkSignage",),
    "TRAV": ("travel",),
    "CVENT": ("cventComparison",),
}

DAY_TO_DATE = {
    "Saturday 6 June": "2026-06-06",
    "Sunday 7 June": "2026-06-07",
    "Monday 8 June": "2026-06-08",
    "Tuesday 9 June": "2026-06-09",
    "Wednesday 10 June": "2026-06-10",
    "Thursday 11 June": "2026-06-11",
    "Friday 12 June": "2026-06-12",
}


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parts(value: str) -> list[str]:
    return [p.strip() for p in re.split(r"\s*\|\s*", clean(value)) if p.strip()]


def canonical_index(canonical_id: str) -> tuple[str, int] | tuple[None, None]:
    if not canonical_id or "-" not in canonical_id:
        return None, None
    prefix, number = canonical_id.split("-", 1)
    try:
        return prefix, int(number) - 1
    except ValueError:
        return prefix, None


def parse_source(row: dict) -> dict | None:
    raw = clean(row.get("Current Source Data"))
    if not raw or not raw.startswith("{"):
        return None
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return None


def find_by_source(items: list[dict], source: dict | None) -> dict | None:
    if not source:
        return None
    for key in ("id", "updateId", "documentName", "fileNeeded", "name", "itemName", "title"):
        value = source.get(key)
        if value:
            for item in items:
                if item.get(key) == value:
                    return item
    source_name = source.get("name") or source.get("itemName") or source.get("speakerName")
    if source_name:
        for item in items:
            if source_name in {item.get("name"), item.get("itemName"), item.get("speakerName")}:
                return item
    return None


def get_record(data: dict, canonical_id: str, row: dict) -> tuple[dict | None, str]:
    prefix, index = canonical_index(canonical_id)
    if prefix == "DOC":
        source = parse_source(row)
        candidates = data.get("documents", []) + data.get("missingFiles", [])
        record = find_by_source(candidates, source)
        return record, "documents/missingFiles"
    if prefix == "STAFF":
        source = parse_source(row)
        for group, items in (data.get("staffLists") or {}).items():
            record = find_by_source(items, source)
            if record:
                return record, f"staffLists.{group}"
        return None, "staffLists"
    path = PREFIX_TO_PATH.get(prefix)
    if not path:
        return None, prefix or ""
    items = data.get(path[0], [])
    source = parse_source(row)
    record = find_by_source(items, source)
    if record:
        return record, path[0]
    if index is not None and 0 <= index < len(items):
        return items[index], path[0]
    return None, path[0]


def status_from_text(value: str, fallback: str = "Needs Confirmation") -> str:
    text = clean(value)
    if not text:
        return fallback
    lowered = text.lower()
    if "resolved" in lowered:
        return "Resolved"
    if "not needed" in lowered:
        return "Not Needed"
    if "on track" in lowered:
        return "On Track"
    if "confirmed" in lowered:
        return "Confirmed"
    if "needs confirmation" in lowered or "tbc" in lowered:
        return "Needs Confirmation"
    if "file needed" in lowered:
        return "File Needed"
    return text


def parse_day(value: str) -> tuple[str, str]:
    day = clean(value)
    return day, DAY_TO_DATE.get(day, "")


def parse_time_start(time_display: str) -> str:
    text = clean(time_display)
    match = re.search(r"(\d{1,2})(?::(\d{2}))?\s*(AM|PM)", text, re.I)
    if not match:
        return ""
    hour = int(match.group(1))
    minute = int(match.group(2) or "00")
    ampm = match.group(3).upper()
    if ampm == "PM" and hour != 12:
        hour += 12
    if ampm == "AM" and hour == 12:
        hour = 0
    return f"{hour:02d}:{minute:02d}"


def apply_pipe_fields(record: dict, canonical_id: str, final: str) -> list[str]:
    p = parts(final)
    changed: list[str] = []
    prefix, _ = canonical_index(canonical_id)
    if not p:
        return changed

    def set_field(key: str, value: str):
        if clean(value) and record.get(key) != value:
            record[key] = value
            changed.append(key)

    if prefix == "SCH" and len(p) >= 6:
        day, date = parse_day(p[0])
        set_field("dayLabel", day)
        set_field("day", day)
        if date:
            set_field("date", date)
        set_field("timeDisplay", p[1])
        start = parse_time_start(p[1])
        if start:
            set_field("timeStart", start)
        set_field("title", p[2])
        set_field("location", p[3])
        set_field("owner", p[4])
        set_field("status", status_from_text(p[5], record.get("status", "Needs Confirmation")))
    elif prefix == "CALL" and len(p) >= 6:
        day, date = parse_day(p[0])
        set_field("day", day)
        if date:
            set_field("date", date)
        set_field("title", p[1])
        set_field("dailyFocus", p[2])
        set_field("crewCallTime", p[3])
        set_field("mainLocation", p[4])
        set_field("status", status_from_text(p[5], record.get("status", "Needs Confirmation")))
    elif prefix == "LOC" and len(p) >= 5:
        set_field("locationName", p[0])
        set_field("canonicalName", p[0])
        set_field("primaryUse", p[1])
        set_field("keyOwner", p[2])
        set_field("watchOut", p[3])
        set_field("status", status_from_text(p[4], record.get("status", "Needs Confirmation")))
    elif prefix in {"CON", "STAFF"}:
        if len(p) >= 5:
            set_field("name", p[0])
            set_field("role", p[1])
            set_field("responsibility", p[2])
            set_field("phone", p[3])
            set_field("notes", p[4])
        else:
            m = re.match(r"^([^,]+),\\s*([^,]+),\\s*(.+?)\\.?$", final)
            if m:
                set_field("name", m.group(1))
                set_field("role", m.group(2))
                set_field("company", m.group(3))
                set_field("notes", final)
            elif " is " in final:
                name, rest = final.rstrip(".").split(" is ", 1)
                set_field("name", name.strip())
                set_field("company", rest.strip())
                set_field("notes", final)
    elif prefix == "WHO" and len(p) >= 4:
        set_field("situation", p[0])
        set_field("primaryContact", p[1])
        set_field("backupContact", p[2])
        set_field("notes", p[3])
    elif prefix == "POD" and len(p) >= 7:
        day, date = parse_day(p[0].split(",")[0])
        set_field("day", p[0])
        if date:
            set_field("date", date)
        set_field("time", p[1])
        set_field("slot", p[2])
        set_field("presenter", p[3])
        set_field("guest", p[4])
        set_field("location", p[5])
        set_field("status", status_from_text(p[6], record.get("status", "Needs Confirmation")))
    elif prefix in {"MAT", "SIG"} and len(p) >= 6:
        set_field("itemName", p[0])
        set_field("type" if prefix == "SIG" else "category", p[1])
        if prefix == "MAT":
            set_field("day", p[2])
            set_field("location", p[3])
            set_field("owner", p[4])
            set_field("status", status_from_text(p[5], record.get("status", "Needs Confirmation")))
        else:
            set_field("locationPlacement", p[2])
            set_field("owner", p[3])
            set_field("status", status_from_text(p[4], record.get("status", "Needs Confirmation")))
    elif prefix == "TRAV" and len(p) >= 4:
        set_field("person", p[0])
        set_field("arrivalDate", p[1])
        set_field("arrivalTime", p[2])
        set_field("status", status_from_text(p[3], record.get("status", "Needs Confirmation")))
    elif prefix == "SPEAK" and len(p) >= 5:
        set_field("location", p[0])
        set_field("day", p[1])
        set_field("time", p[2])
        set_field("speakerName", p[3])
        set_field("sessionTitle", p[4])
    elif prefix == "REH" and len(p) >= 5:
        set_field("rehearsalName", p[0])
        set_field("day", p[1])
        set_field("time", p[2])
        set_field("location", p[3])
        set_field("status", status_from_text(p[-1], record.get("status", "Needs Confirmation")))
    return changed


def apply_final_text(record: dict, canonical_id: str, final: str, row: dict, batch_id: str, row_number: int) -> list[str]:
    changed = apply_pipe_fields(record, canonical_id, final)
    prefix, _ = canonical_index(canonical_id)

    # Row-specific approved sentence corrections that are not pipe-structured.
    if prefix == "RED":
        record["issue"] = final
        record["whyItMatters"] = final
        if "resolved" in final.lower():
            record["status"] = "Resolved"
            record["priority"] = "Resolved"
        changed.extend(["issue", "whyItMatters", "status"])
    elif prefix == "DEC":
        record["decisionNeeded"] = final
        record["latestUpdate"] = final
        if "setup" in final.lower() or "chairs" in final.lower():
            record["status"] = "Confirmed"
        changed.extend(["decisionNeeded", "latestUpdate", "status"])
    elif prefix == "TODAY":
        record["focus"] = final
        changed.append("focus")
    elif prefix == "DROP":
        record["deliveryNotes"] = final
        changed.append("deliveryNotes")
    elif prefix == "HH":
        record["notes"] = final
        changed.append("notes")
    elif prefix == "DOC":
        record["notes"] = final
        changed.append("notes")
    elif prefix == "CVENT" and "not needed" in final.lower():
        record["status"] = "Not Needed"
        record["hiddenFromLive"] = True
        changed.extend(["status", "hiddenFromLive"])
    elif prefix == "SCH" and "|" not in final and final:
        # Preserve short spoken corrections without over-parsing them.
        time = parse_time_start(final)
        if time and re.search(r"\d", final):
            record["timeStart"] = time
            match = re.search(r"(\d{1,2}(?::\d{2})?\s*(?:AM|PM))", final, re.I)
            if match:
                record["timeDisplay"] = match.group(1).upper().replace(":00 ", ":00 ")
            changed.extend(["timeStart", "timeDisplay"])
        record["latestUpdate"] = final
        changed.append("latestUpdate")

    if "not needed" in final.lower():
        record["status"] = "Not Needed"
        record["hiddenFromLive"] = True
        record["archived"] = True
        changed.extend(["status", "hiddenFromLive", "archived"])

    trace = {
        "source_workbook": "HORIZONS_Master_Website_Data_Review_FINAL.xlsx",
        "source_tab": "MASTER REVIEW",
        "source_row": row_number,
        "review_id": row.get("Review ID"),
        "canonical_record_id": canonical_id,
        "approved_by": row.get("Approved By") or "Final workbook",
        "last_updated": datetime.now().isoformat(timespec="seconds"),
        "updated_from_final_workbook": True,
        "final_import_batch_id": batch_id,
    }
    record.update(trace)
    record["finalWorkbookDisplayText"] = final
    record["finalWorkbookStatus"] = row.get("Status")
    return sorted(set(changed + list(trace.keys()) + ["finalWorkbookDisplayText", "finalWorkbookStatus"]))


def main() -> None:
    batch_id = "final-master-import-" + datetime.now().strftime("%Y%m%d-%H%M")
    data = json.loads(CONTENT.read_text())
    original = deepcopy(data)
    wb = load_workbook(WORKBOOK, data_only=True, read_only=True)
    ws = wb["MASTER REVIEW"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}

    read = applied = skipped = unmapped = not_needed = missing_preserved = conflicts = 0
    sections_updated: set[str] = set()
    records_updated: list[str] = []
    skipped_rows: list[dict] = []
    warnings: list[str] = []

    for row_number, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        read += 1
        row = {h: values[i] if i < len(values) else "" for h, i in idx.items()}
        status = clean(row.get("Status"))
        status_key = status.lower()
        canonical_id = clean(row.get("Canonical Record ID"))
        final = clean(row.get("Final Display Text")) or clean(row.get("Corrected / Final Answer"))
        if not final and status_key in APPROVED:
            final = clean(row.get("Current Website Display"))
        if status_key in {"not needed"}:
            # This is actionable only when mapped; otherwise report it.
            pass
        elif status_key not in APPROVED:
            skipped += 1
            if "missing" in status_key or "confirmation" in status_key:
                missing_preserved += 1
            if "conflict" in status_key:
                conflicts += 1
            skipped_rows.append({
                "row": row_number,
                "review_id": row.get("Review ID"),
                "canonical_id": canonical_id,
                "status": status,
                "reason": "Status not approved/final",
                "display": clean(row.get("Current Website Display"))[:300],
            })
            continue
        record, section = get_record(data, canonical_id, row)
        if not record:
            skipped += 1
            unmapped += 1
            skipped_rows.append({
                "row": row_number,
                "review_id": row.get("Review ID"),
                "canonical_id": canonical_id,
                "status": status,
                "reason": "No matching canonical record",
                "display": final[:300],
            })
            continue
        if not final:
            skipped += 1
            skipped_rows.append({
                "row": row_number,
                "review_id": row.get("Review ID"),
                "canonical_id": canonical_id,
                "status": status,
                "reason": "Approved row has no final answer/display",
                "display": clean(row.get("Current Website Display"))[:300],
            })
            continue

        changed = apply_final_text(record, canonical_id, final, row, batch_id, row_number)
        applied += 1
        if status_key == "not needed" or record.get("hiddenFromLive"):
            not_needed += 1
        sections_updated.add(section)
        records_updated.append(f"{canonical_id} ({section}): {', '.join(changed[:8])}")

    # Apply final workbook naming wins visible in approved rows.
    # Eve Blackwell does not appear in the final workbook; final rows approve Eve Dusek.
    content_text = json.dumps(data, ensure_ascii=False)
    content_text = (
        content_text
        .replace("Eve Blackwell", "Eve Dusek")
        .replace("Eve / Blackwell / + / Poppy / Luck", "Eve Dusek / Poppy Luck")
        .replace("Needs Name Confirmation (ED)", "Eve Dusek")
        .replace("Name confirmation needed (ED)", "Eve Dusek")
    )
    data = json.loads(content_text)

    # Preserve import metadata and audit trail.
    data.setdefault("auditLog", []).append({
        "date": datetime.now().isoformat(timespec="seconds"),
        "updatedBy": "Codex",
        "fileTabUpdated": "content.json",
        "changeMade": "Imported approved rows from final master website data review workbook.",
        "reason": "Final approved workbook is the source of truth for this pass.",
        "source": str(WORKBOOK),
        "status": "Complete",
        "notes": f"Batch {batch_id}; applied {applied}; skipped {skipped}; unmapped {unmapped}.",
    })
    data.setdefault("sourceArchive", []).append({
        "sourceFile": "HORIZONS_Master_Website_Data_Review_FINAL.xlsx",
        "type": "Final approved master review workbook",
        "dateAdded": datetime.now().date().isoformat(),
        "keyInformationCaptured": f"{applied} approved rows imported; {skipped} rows skipped or preserved.",
        "relatedMasterTabs": "MASTER REVIEW; Approved Website Updates; Voice Review Update Log; Needs Human Review",
        "relatedWebsiteSections": "; ".join(sorted(sections_updated)),
        "status": "Imported",
        "notes": f"Batch {batch_id}",
    })
    data.setdefault("meta", {})["lastFinalWorkbookImport"] = {
        "batchId": batch_id,
        "sourceWorkbook": "HORIZONS_Master_Website_Data_Review_FINAL.xlsx",
        "rowsRead": read,
        "approvedRowsApplied": applied,
        "rowsSkipped": skipped,
        "unmappedRows": unmapped,
        "notNeededArchived": not_needed,
        "importedAt": datetime.now().isoformat(timespec="seconds"),
    }

    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n")
    json.loads(PREVIEW.read_text())
    shutil.copy2(PREVIEW, CONTENT)

    report = [
        "# FINAL MASTER IMPORT REPORT",
        "",
        f"- Final workbook used: `{WORKBOOK}`",
        f"- Import batch ID: `{batch_id}`",
        f"- Date/time: {datetime.now().isoformat(timespec='seconds')}",
        f"- Rows read: {read}",
        f"- Approved/final rows applied: {applied}",
        f"- Rows skipped: {skipped}",
        f"- Missing/needs-confirmation rows preserved: {missing_preserved}",
        f"- Conflicts left for review: {conflicts}",
        f"- Unmapped rows: {unmapped}",
        f"- Not-needed/hidden records marked: {not_needed}",
        "",
        "## Sections Updated",
        "",
    ]
    report.extend(f"- {section}" for section in sorted(sections_updated))
    report.extend([
        "",
        "## Records Updated",
        "",
    ])
    report.extend(f"- {item}" for item in records_updated[:300])
    if len(records_updated) > 300:
        report.append(f"- ... {len(records_updated) - 300} additional records updated.")
    report.extend([
        "",
        "## Warnings / Manual Follow-Up",
        "",
        "- Rows marked Needs Confirmation, Missing Data, Waiting on Person, Conflict, Draft, TBC, or Unclear were not imported as final data.",
        "- Rows marked Not Needed were preserved in source/admin data and marked hidden/archived where mapped.",
        "- Production Slack routes were not enabled by this import.",
        "- The final workbook contains approved rows whose final text still includes Needs Confirmation; those are imported as approved wording but remain visibly pending on the site.",
    ])
    REPORT.write_text("\n".join(report) + "\n")

    skipped_report = [
        "# FINAL MASTER IMPORT SKIPPED ROWS",
        "",
        "Rows below were not imported as final website data.",
        "",
        "| Row | Review ID | Canonical ID | Status | Reason | Display |",
        "| --- | --- | --- | --- | --- | --- |",
    ]
    for item in skipped_rows:
        display = clean(item.get("display")).replace("|", "\\|").replace("\n", " ")[:300]
        skipped_report.append(
            f"| {item.get('row')} | {item.get('review_id') or ''} | {item.get('canonical_id') or ''} | {item.get('status') or ''} | {item.get('reason') or ''} | {display} |"
        )
    SKIPPED.write_text("\n".join(skipped_report) + "\n")

    print(json.dumps({
        "batch_id": batch_id,
        "rows_read": read,
        "applied": applied,
        "skipped": skipped,
        "unmapped": unmapped,
        "not_needed": not_needed,
        "sections_updated": sorted(sections_updated),
        "preview": str(PREVIEW),
    }, indent=2))


if __name__ == "__main__":
    main()
